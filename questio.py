#!/usr/bin/env python3



import os
import datetime
from openpyxl import Workbook


cwd = os.getcwd()

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted

ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
